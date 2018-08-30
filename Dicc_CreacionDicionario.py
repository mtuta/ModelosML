## -*- coding: utf-8 -*- 
# ================================DESCRIPCION===============================
# Código que permite generar un diccionario de palabras de acuerdo a las 
# descripciones de actividades registradas en ARES
# ==================================INPUTS==================================
# 1. Archivo excel con descripciones. El nombre de este archivo debe ser
#   Dicc_Descripciones.xlsx, la pestaña con las descripciones debe llamarse
#   DESCRIPCIONES y las descripciones deben estar en la columna A
# 2 Cantidad de palabras que tendrá el diccionario  (El usuario lo ingresa)
# =================================OUTPUTS==================================
# 1. Archivo .txt con las palabras más repetidas en el excel de entrada Cambio

import string
from openpyxl import load_workbook
import unicodedata
import re

def main():
    #Cantidad de palabras que se quieren obtener en el diccionario
	cant_palabrasDic = input("Ingrese el numero de palabras para el diccionario: ")
	cant_palabrasDic=int(cant_palabrasDic)

    #Se abre/crea el archivo que va a contener el diccionario de palabras. La 'w'
    #indica que el archivo se va a escribir (write)
	file = open('Dicc_ListaPalabras.txt', 'w')
    
    #Dic estructura de datos que mapea las palabras contrab la cantidad de veces que aparecen  
	dic = {}

	numeros = ['0','1','2','3','4','5','6','7','8','9']

	palabrasEliminar = ['al',	'ante',	'bajo',	'cabe',	'como',	'con',	'contra',	'de',	'del',	'desde',	'donde',	'durante',	'el',	'en',
                     'entre',	'es',	'esto',	'hacia',	'hasta',	'hay',	'la',	'las',	'lo',	'los',	'mas',	'mediante',	'no',	'para',
                     'por',	'que',	'se',	'segun',	'si',	'sin',	'sobre',	'sus',	'tras',	'un',	'una',	'unas',	'unos',	'ya']


    #Se lee el archivo excel que contiene las actividades
	wb = load_workbook('Dicc_Descripciones.xlsx')
	sheet_ranges = wb['DESCRIPCIONES']
    #Cantidad de filas en el excel 
	row_count = sheet_ranges.max_row
    

	for i in range(1,row_count):
        # traigo el valor de la columna A con fila i
		actual = sheet_ranges['A' + str(i)].value
        #TrasformaciÃ³n de la cadena para quitar tildes y simbolos del lenguaje español
		if actual is not None:
			actual = unicodedata.normalize('NFD', actual)
			actual = actual.encode('utf8').decode('ascii', 'ignore')
		
			#string.puntuation es un arreglo donde se encuentran todos los signos de puntuaciÃ³n
			#reemplaza todos los signos de puntuacion por doble espacio
			for c in string.punctuation:
				actual = actual.replace(c,'  ')

			#reemplaza los digitos del 0 al 9 por doble espacio
			for n in numeros:
				actual = actual.replace(n,'  ')

			#Reduce todos los espacios para que sea de longitud 1
			#Todos los espacios entre poalabras serÃ¡s espacios simples
			while '  ' in actual:
				actual = actual.replace('  ', ' ')

			#Elimina espacios antes y despues de toda la cadena completa
			actual = actual.strip()
			#Pasa todo a minÃºscula
			actual = actual.lower()
          #ExpresiÃ³n regular que quita todos los cambios de linea...
			actual = re.sub('\s+',' ',actual)

			#arreglo con todas las palabras de la cadena
			listaActual = actual.split(' ')
			#Recorrer todas las palabras dentro de la descripciÃ³n
			for j in range(len(listaActual)):
				# a es la palabra en la posicion j de la cadena actual(arreglo de palabras de la descripciÃ³n)
				a = listaActual[j]
     
				#No se toman vacios ni palabras de una letra
				if a != '' and len(a) > 1:
                 #Si la palabra es nueva la agrego al diccionario y le asigno 1
					#En el diccionario los indices son las palabras y los valores las frecuencias
					if  a not in dic:
						dic[a] = 1
                #La palabra ya existe en el diccionario
					else:
						dic[a] = dic[a] + 1

	# El diccionario se convierte a lista con dic.items() y se organiza segun la frecuencia
	# en orden descendente (reverse)					
	sorted_by_value = sorted(dic.items(), key=lambda kv: kv[1], reverse=True)
	w=0
	p=0
    
	while p <cant_palabrasDic:					
		al = sorted_by_value[w]
		encontro = False
		e=0
      #Se evalua si la palabra de la lista-diccionario se encuentra en el arreglo
      #de palabras a eliminar. Si es asÃ­, no se tiene en cuenta para guardar en
      #archivo file
		while e < len(palabrasEliminar) and not encontro:
				if al[0]==palabrasEliminar[e]:
						encontro=True
				e=e+1
      #Se guarda palabra en diccionario si: *Está entre las más frecuentes, *No estaba antes
      #en el arreglo de palabaras a eliminar
		if not encontro:
				file.write(str(p+1)+' '+al[0] +' '+ str(al[1])+'\n') 
				p=p+1
		w=w+1
    
	file.close()
	print('===================FIN DE LA OPERACION==================')
# Ejecución de la función
if __name__ == '__main__':
		main()

	
